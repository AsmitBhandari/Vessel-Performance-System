import io
import datetime
from typing import Dict, Any, List, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_OPERATIONAL_LABELS = {"utcdate", "latitude", "longitude"}

# Centralized label alias mapping
# Map normalized search string -> (JSON_group_key, JSON_field_key)
# If JSON_group_key is None, it is placed at the root of the report object.
ALIAS_MAP: Dict[str, Tuple[Optional[str], str]] = {
    # Metadata
    "vessel": (None, "vesselName"),
    "vesselname": (None, "vesselName"),
    "techmanager": (None, "technicalManager"),
    "technicalmanager": (None, "technicalManager"),
    
    # Root level report fields
    "utcdate": (None, "utcDate"),
    "utctime": (None, "utcTime"),
    "vesselcondition": (None, "vesselCondition"),
    "otherremarksifany": (None, "remarks"),
    
    # Position
    "latitude": ("position", "latitude"),
    "longitude": ("position", "longitude"),
    
    # Voyage
    "draughtfwd": ("voyage", "draftForward"),
    "draughtaft": ("voyage", "draftAft"),
    "lastport": ("voyage", "lastPort"),
    "cospdate": ("voyage", "cospDate"),
    "cosptime": ("voyage", "cospTime"),
    "nextport": ("voyage", "nextPort"),
    "distancetonextport": ("voyage", "distanceToNextPort"),
    "distancesailed": ("voyage", "distanceSailed"),
    "enginedistance24hrs": ("voyage", "engineDistance"),
    "distancesailedfromcosp": ("voyage", "distanceSailedFromCosp"),
    "steamingtimedaily": ("voyage", "steamingTimeDaily"),
    "steamingtimeaftercosp": ("voyage", "steamingTimeAfterCosp"),
    "allowedcpspeed": ("voyage", "allowedCpSpeed"),
    "speedlast24hrs": ("voyage", "speedLast24Hrs"),
    "avgspeed": ("voyage", "avgSpeed"),
    "mainenginerpm": ("voyage", "meRpm"),
    "averageslip": ("voyage", "avgSlip"),
    "etadatenextport": ("voyage", "etaDate"),
    "etatimenextport": ("voyage", "etaTime"),
    
    # Weather
    "windspeed": ("weather", "windSpeed"),
    "winddirection": ("weather", "windDirection"),
    "currentdirection": ("weather", "currentDirection"),
    "buefortscale": ("weather", "beaufortScale"),
    "beaufortscale": ("weather", "beaufortScale"),
    
    # Lube Oil
    "mecyloilrob": ("lubeOil", "meCylOilRob"),
    "cyloilconsumption": ("lubeOil", "cylOilConsumption"),
    "mesystemoilrob": ("lubeOil", "meSystemOilRob"),
    "consumptionmesystemoil": ("lubeOil", "meSystemOilConsumption"),
    "robaelo": ("lubeOil", "aeLoRob"),
    "consumptionaelo": ("lubeOil", "aeLoConsumption"),
    
    # Bunker Consumption
    "mehsfoconsumptionmt": ("bunker", "meHsfoConsumption"),
    "mehsfoconsumption": ("bunker", "meHsfoConsumption"),
    "melsfoconsumptionmt": ("bunker", "meLsfoConsumption"),
    "melsfoconsumption": ("bunker", "meLsfoConsumption"),
    "memgoconsumptionmt": ("bunker", "meMgoConsumption"),
    "memgoconsumption": ("bunker", "meMgoConsumption"),
    "aehfoconsumptionmt": ("bunker", "aeHsfoConsumption"),
    "aehfoconsumption": ("bunker", "aeHsfoConsumption"),
    "aehsfoconsumptionmt": ("bunker", "aeHsfoConsumption"),
    "aehsfoconsumption": ("bunker", "aeHsfoConsumption"),
    "aelsfoconsumptionmt": ("bunker", "aeLsfoConsumption"),
    "aelsfoconsumption": ("bunker", "aeLsfoConsumption"),
    "aemgoconsumptionmt": ("bunker", "aeMgoConsumption"),
    "aemgoconsumption": ("bunker", "aeMgoConsumption"),
    "boilerhsfoconsumtionmt": ("bunker", "boilerHsfoConsumption"),
    "boilerhsfoconsumtion": ("bunker", "boilerHsfoConsumption"),
    "boilerlsfoconsumptionmt": ("bunker", "boilerLsfoConsumption"),
    "boilerlsfoconsumption": ("bunker", "boilerLsfoConsumption"),
    "boilermgoconsumptionmt": ("bunker", "boilerMgoConsumption"),
    "boilermgoconsumption": ("bunker", "boilerMgoConsumption"),
    "totalhsfoconsumption": ("bunker", "totalHsfoConsumption"),
    "totallsfoconsumption": ("bunker", "totalLsfoConsumption"),
    "totalmgoconsumption": ("bunker", "totalMgoConsumption"),
    "hsforeceivedmt": ("bunker", "hsfoReceived"),
    "hsforeceived": ("bunker", "hsfoReceived"),
    "lsforeceivedmt": ("bunker", "lsfoReceived"),
    "lsforeceived": ("bunker", "lsfoReceived"),
    "mgoreceivedmt": ("bunker", "mgoReceived"),
    "mgoreceived": ("bunker", "mgoReceived"),
    "robhsfomt": ("bunker", "hsfoRob"),
    "robhsfo": ("bunker", "hsfoRob"),
    "roblsfomt": ("bunker", "lsfoRob"),
    "roblsfo": ("bunker", "lsfoRob"),
    "robmgomt": ("bunker", "mgoRob"),
    "robmgo": ("bunker", "mgoRob"),
    
    # Fresh Water
    "freshwatergeneration": ("freshWater", "fwGeneration"),
    "freshwaterconsumed": ("freshWater", "fwConsumed"),
    "freshwaterreceived": ("freshWater", "fwReceived"),
    "freshwaterrob": ("freshWater", "fwRob"),
    
    # Machinery
    "ae1runninghours": ("machinery", "ae1RunningHours"),
    "ae2runninghours": ("machinery", "ae2RunningHours"),
    "ae3runninghours": ("machinery", "ae3RunningHours"),
}

def standardize_label(label: Optional[str]) -> str:
    """Clean and normalize label text for resilient anchor matching."""
    if label is None:
        return ""
    return "".join(c for c in str(label).lower() if c.isalnum())

def discover_row_anchors(sheet: Worksheet) -> Dict[str, int]:
    """
    Scan Column A and Column B of the worksheet.
    Map standardized labels to their row indices.
    """
    anchors = {}
    
    # Scan Column A (Col 1) and Column B (Col 2) for label row anchors
    for r in range(1, min(sheet.max_row + 1, 200)):  # Scan first 200 rows
        col_a_val = sheet.cell(row=r, column=1).value
        col_b_val = sheet.cell(row=r, column=2).value
        
        if col_a_val is not None:
            norm_a = standardize_label(str(col_a_val))
            if norm_a in ALIAS_MAP:
                anchors[norm_a] = r
                
        if col_b_val is not None:
            norm_b = standardize_label(str(col_b_val))
            if norm_b in ALIAS_MAP:
                # Store the most specific/applicable alias key
                anchors[norm_b] = r
                
    return anchors

def coerce_float(val: Any) -> Optional[float]:
    """Safely convert Excel cell values into Floats or None."""
    if val is None:
        return None
    val_str = str(val).strip().upper()
    if val_str in {"N/A", "NIL", "-", "N.A.", "N/D", "*" , ""}:
        return None
    try:
        return float(val)
    except ValueError:
        return None

def format_date_value(val: Any) -> Optional[str]:
    """Coerce date and datetime values into clean ISO-8601 strings."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%dT%H:%M:%SZ") if isinstance(val, datetime.datetime) else val.strftime("%Y-%m-%d")
    
    val_str = str(val).strip()
    if val_str.upper() in {"N/A", "NIL", "-", "N.A.", ""}:
        return None
    return val_str

def format_string_value(val: Any) -> Optional[str]:
    """Strip string values and handle placeholder nulls."""
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str.upper() in {"N/A", "NIL", "-", "N.A.", ""}:
        return None
    return val_str

def parse_single_sheet(sheet: Worksheet) -> Optional[Tuple[Dict[str, Any], Dict[str, Any]]]:
    """
    Attempt to parse a single worksheet.
    Returns a tuple (parsed_data, parser_info) if it is a valid operational sheet, otherwise None.
    """
    anchors = discover_row_anchors(sheet)
    
    # Check for the required operational labels in Column B
    missing = REQUIRED_OPERATIONAL_LABELS - set(anchors.keys())
    if missing:
        return None
        
    # 1. Extract Vessel Metadata
    vessel_name = None
    vessel_row = anchors.get("vessel") or anchors.get("vesselname")
    if vessel_row:
        vessel_name = format_string_value(sheet.cell(row=vessel_row, column=2).value)
        
    tech_manager = None
    tech_row = anchors.get("techmanager") or anchors.get("technicalmanager")
    if tech_row:
        tech_manager = format_string_value(sheet.cell(row=tech_row, column=2).value)
        
    # 2. Extract Daily Position Reports
    reports = []
    date_row = anchors["utcdate"]
    
    # Iterate columns starting from Column 3
    for c in range(3, sheet.max_column + 1):
        date_cell_val = sheet.cell(row=date_row, column=c).value
        if date_cell_val is None:
            break
        formatted_date = format_date_value(date_cell_val)
        if formatted_date is None:
            break
            
        # Compile structured nested groupings
        report_entry = {
            "position": {},
            "voyage": {},
            "weather": {},
            "lubeOil": {},
            "bunker": {},
            "freshWater": {},
            "machinery": {}
        }
        
        # Populate all registered fields based on anchors
        for alias_key, row_idx in anchors.items():
            # Skip metadata keys during daily report column loop
            if alias_key in {"vessel", "vesselname", "techmanager", "technicalmanager"}:
                continue
                
            group_key, field_key = ALIAS_MAP[alias_key]
            raw_val = sheet.cell(row=row_idx, column=c).value
            
            # Format/Coerce values based on their targets
            if field_key in {"utcDate", "cospDate", "etaDate"}:
                coerced = format_date_value(raw_val)
            elif field_key in {
                "utcTime", "vesselCondition", "lastPort", "cospTime", 
                "nextPort", "etaTime", "latitude", "longitude", "remarks"
            }:
                coerced = format_string_value(raw_val)
            else:
                coerced = coerce_float(raw_val)
                
            # Place in appropriate nesting level
            if group_key is None:
                report_entry[field_key] = coerced
            else:
                report_entry[group_key][field_key] = coerced
                
        reports.append(report_entry)
        
    parsed_data = {
        "vesselName": vessel_name or "Unknown",
        "technicalManager": tech_manager or "Unknown",
        "reportCount": len(reports),
        "reports": reports,
    }
    
    parser_info = {
        "sheetName": sheet.title,
        "detectedReportColumns": len(reports),
        "parserVersion": "2.0"
    }
    
    return parsed_data, parser_info

def parse_excel_report(file_stream: io.BytesIO) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Load workbook in memory, identify the best operational sheet candidate,
    and return the parsed dataset and parser metadata.
    """
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    candidates = []
    
    # Scan all worksheets in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        try:
            res = parse_single_sheet(sheet)
            if res is not None:
                candidates.append(res)
        except Exception:
            continue
            
    if not candidates:
        raise ValueError(
            "No operational noon report sheet detected. "
            "Column B must contain 'UTC DATE', 'LATITUDE', and 'LONGITUDE'."
        )
        
    # Rank candidates by:
    # 1. Primary: Highest number of parsed report columns
    # 2. Secondary: Whether a non-empty vessel name was extracted
    candidates.sort(
        key=lambda x: (
            x[0]["reportCount"], 
            1 if x[0]["vesselName"] != "Unknown" else 0
        ),
        reverse=True
    )
    
    return candidates[0]
